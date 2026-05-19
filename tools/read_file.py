from langchain.tools import tool


def _read_text(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def _read_pdf(path: str) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return "错误: 需要安装 pypdf (pip install pypdf)"
    reader = PdfReader(path)
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _read_xlsx(path: str) -> str:
    try:
        import openpyxl
    except ImportError:
        return "错误: 需要安装 openpyxl (pip install openpyxl)"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            lines.append("\t".join(str(c) if c is not None else "" for c in row))
    return "\n".join(lines)


_READERS = {
    ".txt": _read_text,
    ".py": _read_text,
    ".md": _read_text,
    ".json": _read_text,
    ".yaml": _read_text,
    ".yml": _read_text,
    ".xml": _read_text,
    ".html": _read_text,
    ".css": _read_text,
    ".js": _read_text,
    ".csv": _read_text,
    ".toml": _read_text,
    ".ini": _read_text,
    ".cfg": _read_text,
    ".env": _read_text,
    ".pdf": _read_pdf,
    ".xlsx": _read_xlsx,
}


@tool
def read_file(path: str) -> str:
    """读取文件内容，支持 txt/pdf/xlsx 等格式"""
    import os

    path = os.path.abspath(path)
    if not os.path.exists(path):
        return f"错误: 文件不存在 - {path}"
    if not os.path.isfile(path):
        return f"错误: 不是文件 - {path}"

    ext = os.path.splitext(path)[1].lower()
    reader = _READERS.get(ext)
    if reader is None:
        # 未知格式，尝试按文本读取
        try:
            return _read_text(path)
        except UnicodeDecodeError:
            return f"错误: 不支持的文件格式 '{ext}'，或文件是二进制文件"

    try:
        return reader(path)
    except Exception as e:
        return f"错误: 读取失败 - {e}"

